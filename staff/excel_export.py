from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.http import HttpResponse
from datetime import datetime

def export_payroll_excel(shifts, filename="payroll_report"):
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll"
    
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=11)
    total_fill = PatternFill(start_color="27ae60", end_color="27ae60", fill_type="solid")
    total_font = Font(color="FFFFFF", bold=True, size=12)
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    headers = ['Ξενοδοχείο', 'Τμήμα', 'Υπάλληλος', 'Ρόλος', 'Ημερομηνία', 'Ώρες', 'Ωρομίσθιο', 'Ημερομίσθιο', 'Υπερωρίες', 'Σύνολο']
    ws.append(headers)
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
        cell.border = border
    
    total_wages = 0
    total_overtime = 0
    
    for shift in shifts:
        overtime_amount = sum(float(ot.amount) for ot in shift.overtimes.all())
        total = float(shift.daily_wage) + overtime_amount
        
        row = [
            shift.department.hotel.name,
            shift.department.name,
            shift.employee.full_name,
            shift.employee.role,
            shift.date.strftime('%d/%m/%Y'),
            shift.actual_hours or shift.scheduled_hours,
            float(shift.employee.hourly_wage),
            float(shift.daily_wage),
            overtime_amount,
            total
        ]
        ws.append(row)
        total_wages += float(shift.daily_wage)
        total_overtime += overtime_amount
    
    total_row = [''] * 7 + [total_wages, total_overtime, total_wages + total_overtime]
    ws.append(total_row)
    for cell in ws[ws.max_row]:
        cell.fill = total_fill
        cell.font = total_font
        cell.border = border
    
    for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
        ws.column_dimensions[col].width = 15
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename={filename}_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response

def export_employee_list(employees):
    wb = Workbook()
    ws = wb.active
    ws.title = "Employees"
    
    headers = ['ID', 'Επίθετο', 'Όνομα', 'Ξενοδοχείο', 'Τμήμα', 'Ρόλος', 'Τύπος Σύμβασης', 'Ωρομίσθιο', 'Email', 'Τηλέφωνο', 'Ενεργός']
    ws.append(headers)
    
    for emp in employees:
        ws.append([
            emp.id, emp.last_name, emp.first_name, emp.hotel.name, emp.department.name,
            emp.role, emp.get_contract_type_display(), float(emp.hourly_wage),
            emp.email, emp.phone, 'Ναι' if emp.is_active else 'Όχι'
        ])
    
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=employees_{datetime.now().strftime("%Y%m%d")}.xlsx'
    wb.save(response)
    return response